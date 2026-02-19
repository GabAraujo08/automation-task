import streamlit as st
import holidays
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta
import io
import pandas as pd


# ─── LÓGICA DE DATAS ──────────────────────────────────────────────────────────
FERIADOS_BR = holidays.country_holidays("BR")


def eh_dia_util(d):
    return d.weekday() < 5 and d not in FERIADOS_BR


def proximo_dia_util(d):
    while not eh_dia_util(d):
        d += timedelta(days=1)
    return d


def adicionar_dias_uteis(inicio, n):
    atual, cont = inicio, 0
    while cont < n:
        atual += timedelta(days=1)
        if eh_dia_util(atual):
            cont += 1
    return atual


# ✅ CORREÇÃO 3: Retorna sempre 2 janelas de admissão (segunda-feira, até dia 20)
def janelas_admissao(ref):
    """
    Retorna duas opções de segunda-feira com dia <= 20.
    Se o mês de 'ref' tiver pelo menos uma, usa ela como primeira e busca a segunda.
    Caso contrário, busca as duas primeiras no próximo mês.
    """
    def segundas_ate_dia_20_do_mes(ano, mes):
        resultado = []
        d = date(ano, mes, 1)
        while d.month == mes:
            if d.weekday() == 0 and d.day <= 20:
                resultado.append(d)
            d += timedelta(days=1)
        return resultado

    ref = proximo_dia_util(ref)

    # Coleta segundas válidas no mês atual a partir de ref
    candidatas_mes_atual = [
        d for d in segundas_ate_dia_20_do_mes(ref.year, ref.month)
        if d >= ref
    ]

    # Coleta segundas válidas no próximo mês
    mes_prox = ref.month % 12 + 1
    ano_prox = ref.year + (1 if ref.month == 12 else 0)
    candidatas_prox_mes = segundas_ate_dia_20_do_mes(ano_prox, mes_prox)

    todas = candidatas_mes_atual + candidatas_prox_mes

    # Garante pelo menos 2 opções
    return todas[0], todas[1]


def fmt(d):
    return d.strftime("%d/%m")


def fmt_periodo(ini, fim):
    return fmt(ini) if ini == fim else f"{fmt(ini)} a {fmt(fim)}"


# ─── GERAÇÃO DO EXCEL EM MEMÓRIA ──────────────────────────────────────────────
def gerar_excel(data_rc, data_alinhamento, data_divulgacao_inicio, data_divulgacao_fim):
    etapas = []

    etapas.append(("Recebimento de RC", data_rc, data_rc, "—"))
    etapas.append(("Alinhamento de Perfil", data_alinhamento, data_alinhamento, 1))
    etapas.append(("Divulgação da Vaga", data_divulgacao_inicio, data_divulgacao_fim, 7))

    # ✅ CORREÇÃO 1: Triagem dos Inscritos = 2 dias úteis
    ini = proximo_dia_util(data_divulgacao_fim + timedelta(days=1))
    fim = adicionar_dias_uteis(ini, 1)  # ini + 1 dia útil = 2 dias úteis no total
    etapas.append(("Triagem dos Inscritos", ini, fim, 2))

    # ✅ CORREÇÃO 1: Mapeamento dos Candidatos = 2 dias úteis
    ini = proximo_dia_util(fim + timedelta(days=1))
    fim = adicionar_dias_uteis(ini, 1)
    etapas.append(("Mapeamento dos Candidatos", ini, fim, 2))

    ini = proximo_dia_util(fim + timedelta(days=1))
    fim = adicionar_dias_uteis(ini, 1)
    etapas.append(("Entrevistas RH", ini, fim, 2))

    ini = proximo_dia_util(fim + timedelta(days=1))
    fim = adicionar_dias_uteis(ini, 1)
    etapas.append(("Apresentação Relatório Gestor", ini, fim, 2))

    ini = proximo_dia_util(fim + timedelta(days=1))
    fim = adicionar_dias_uteis(ini, 1)
    etapas.append(("Entrevista Gestor + RH", ini, fim, 2))

    ini = proximo_dia_util(fim + timedelta(days=1))
    fim = adicionar_dias_uteis(ini, 14)
    etapas.append(("Proposta + Processo de Admissão", ini, fim, "10 / 15"))

    # ✅ CORREÇÃO 3: Sempre 2 janelas de admissão garantidas
    ref = proximo_dia_util(fim + timedelta(days=1))
    opcao1, opcao2 = janelas_admissao(ref)
    label = f"{fmt(opcao1)} ou {fmt(opcao2)}"
    etapas.append(("Previsão de Início", opcao1, opcao1, label))

    # ─── EXCEL ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cronograma"

    h_fill  = PatternFill("solid", fgColor="1F4E79")
    h_font  = Font(bold=True, color="FFFFFF", size=11)
    alt     = PatternFill("solid", fgColor="D6E4F0")
    normal  = PatternFill("solid", fgColor="FFFFFF")
    div_fill= PatternFill("solid", fgColor="2E75B6")
    div_font= Font(bold=True, color="FFFFFF", size=10)
    center  = Alignment(horizontal="center", vertical="center")
    left    = Alignment(horizontal="left",   vertical="center")
    borda   = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"),  bottom=Side(style="thin"))

    for col, (h, w) in enumerate(zip(["Atividades","Período","Dias Úteis"],[35,22,12]), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = h_font, h_fill, center, borda
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for i, (ativ, ini, fim, du) in enumerate(etapas, 2):
        fill = div_fill if ativ == "Divulgação da Vaga" else (alt if i%2==0 else normal)
        font = div_font if ativ == "Divulgação da Vaga" else Font(size=10)

        periodo = du if ativ == "Previsão de Início" else fmt_periodo(ini, fim)

        for col, val in enumerate([ativ, periodo, "" if ativ == "Previsão de Início" else str(du)], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill, c.font, c.border = fill, font, borda
            c.alignment = left if col == 1 else center
        ws.row_dimensions[i].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, etapas


# ─── INTERFACE STREAMLIT ──────────────────────────────────────────────────────
st.set_page_config(page_title="Cronograma R&S", page_icon="📋", layout="centered")

st.title("📋 Gerador de Cronograma — R&S")
st.markdown("Preencha as datas e baixe a planilha pronta.")

col1, col2 = st.columns(2)
with col1:
    data_rc = st.date_input("📥 Recebimento do RC", value=date.today(), format="DD/MM/YYYY", key="rc")
with col2:
    data_ali = st.date_input("🎯 Alinhamento de Perfil", value=date.today(), format="DD/MM/YYYY", key="ali")

st.divider()

# ✅ CORREÇÃO 2: Apenas data início da divulgação; fim calculado automaticamente (15 dias úteis)
data_div_ini = st.date_input("📢 Divulgação — Data Início", value=date.today(), format="DD/MM/YYYY", key="div_ini")
data_div_fim = adicionar_dias_uteis(data_div_ini, 15)
st.info(f"📅 Data fim da divulgação calculada automaticamente: **{data_div_fim.strftime('%d/%m/%Y')}** (15 dias úteis)")

if st.button("⚡ Gerar Cronograma", use_container_width=True, type="primary"):
    buf, etapas = gerar_excel(data_rc, data_ali, data_div_ini, data_div_fim)

    st.success("✅ Cronograma gerado com sucesso!")

    # ─── TABELA LIMPA PARA PRINT ─────────────────────────────────────────
    st.markdown("### 📅 Cronograma")

    dados_tabela = []
    for ativ, ini, fim, du in etapas:
        if ativ == "Previsão de Início":
            periodo = du
            dias_util = "—"
        else:
            periodo = fmt_periodo(ini, fim)
            dias_util = str(du)

        dados_tabela.append({
            "Atividades": ativ,
            "Período": periodo,
            "Dias Úteis": dias_util
        })

    df = pd.DataFrame(dados_tabela)

    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Atividades": st.column_config.TextColumn(width="large"),
            "Período": st.column_config.TextColumn(width="medium"),
            "Dias Úteis": st.column_config.TextColumn(width="small"),
        }
    )

    st.markdown("*Você pode tirar um print dessa tabela*")
    st.divider()

    # ─── DOWNLOAD DO EXCEL ───────────────────────────────────────────────
    nome = f"cronograma_{data_rc.strftime('%d%m%Y')}.xlsx"
    st.download_button(
        label="⬇️ Baixar planilha Excel",
        data=buf,
        file_name=nome,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
